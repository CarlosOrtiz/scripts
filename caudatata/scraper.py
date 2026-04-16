"""
Scraper de tirillas para Caudata.
Uso recomendado:
    - BASE_URL="https:// tu dominio de la empresa .caudata.me" ej. "https://caol.caudata.me"
    - CAUDATA_DOCUMENT="document"
    - CAUDATA_PASSWORD="password"
    - python caudata_scraper.py --output caudata_pagos.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from typing import Dict, List, Optional
from urllib.parse import urljoin

try:
    from curl_cffi import requests as http_requests  # type: ignore

    USING_CURL_CFFI = True
except Exception:
    import requests as http_requests  # type: ignore

    USING_CURL_CFFI = False

from bs4 import BeautifulSoup  # type: ignore
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_URL = "https://*****.caudata.me"
LOGIN_URL = BASE_URL + "/zonatrabajador/index"
TIRILLA_INDEX_URL = BASE_URL + "/zonatrabajador/tirilla/index"

# Puedes cambiar estos valores si quieres
CAUDATA_DOCUMENT = "*"
CAUDATA_PASSWORD = "*"


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "")).strip()


def money_to_int(value: str) -> Optional[int]:
    value = clean_text(value)
    if not value:
        return None

    value = value.replace("$", "").replace(",", "").replace(".", "")
    value = re.sub(r"[^\d-]", "", value)

    if not value:
        return None

    try:
        return int(value)
    except ValueError:
        return None


def format_currency(cell) -> None:
    cell.number_format = "$#,##0;[Red]($#,##0)"


def session_factory():
    if USING_CURL_CFFI:
        try:
            return http_requests.Session(impersonate="chrome124")
        except TypeError:
            return http_requests.Session()
    return http_requests.Session()


def absolute_url(url: str) -> str:
    return urljoin(BASE_URL, url)


@dataclass
class TirillaRow:
    orden: int
    fecha_pago: str
    mes: str
    request_path: str
    id_tirilla: str
    periodo_pago: Optional[str] = None
    salario_basico: Optional[int] = None
    total_devengos: Optional[int] = None
    total_deducciones: Optional[int] = None
    total_a_pagar: Optional[int] = None
    fuente: Optional[str] = None


class ZingScraper:
    def __init__(
        self,
        document: str,
        password: str,
        timeout: int = 30,
        verbose: bool = False,
    ):
        self.document = document
        self.password = password
        self.timeout = timeout
        self.verbose = verbose
        self.session = session_factory()

        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0 Safari/537.36"
                ),
                "Accept": (
                    "text/html,application/xhtml+xml,application/xml;q=0.9,"
                    "image/avif,image/webp,image/apng,*/*;q=0.8"
                ),
                "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
                "Connection": "keep-alive",
            }
        )

    def log(self, *args) -> None:
        if self.verbose:
            print(*args, file=sys.stderr)

    def get(self, url: str):
        resp = self.session.get(url, timeout=self.timeout, allow_redirects=True)
        resp.raise_for_status()
        return resp

    def post(self, url: str, data: Dict[str, str], referer: Optional[str] = None):
        headers = {}
        if referer:
            headers["Referer"] = referer

        resp = self.session.post(
            url,
            data=data,
            headers=headers,
            timeout=self.timeout,
            allow_redirects=True,
        )
        resp.raise_for_status()
        return resp

    def find_form_payload(self, soup: BeautifulSoup, input_id: str, value: str):
        input_el = soup.select_one(f"#{input_id}")
        if not input_el:
            return None, None

        form = input_el.find_parent("form")
        if not form:
            return None, None

        action = form.get("action") or ""
        payload: Dict[str, str] = {}

        for elem in form.select("input, select, textarea"):
            name = elem.get("name")
            if not name:
                continue

            tag = elem.name.lower()
            elem_type = (elem.get("type") or "").lower()

            if tag == "input":
                if elem_type in {"checkbox", "radio"} and not elem.has_attr("checked"):
                    continue
                payload[name] = elem.get("value", "")
            elif tag == "textarea":
                payload[name] = elem.text or ""
            elif tag == "select":
                selected = elem.select_one("option[selected]") or elem.select_one(
                    "option"
                )
                payload[name] = selected.get("value", "") if selected else ""

        field_name = input_el.get("name") or input_id
        payload[field_name] = value

        # Algunos formularios usan botón submit con name/value
        submit_btn = form.select_one(
            'button[type="submit"][name], input[type="submit"][name]'
        )
        if submit_btn:
            submit_name = submit_btn.get("name")
            submit_value = submit_btn.get("value", "")
            if submit_name:
                payload[submit_name] = submit_value

        return action, payload

    def ensure_logged_in(self) -> None:
        if not self.document:
            raise RuntimeError("No hay documento.")
        if not self.password:
            raise RuntimeError("No hay password.")

        # Paso 1: cargar login
        resp1 = self.get(LOGIN_URL)
        soup1 = BeautifulSoup(resp1.text, "html.parser")

        action1, payload1 = self.find_form_payload(soup1, "asodocumento", self.document)
        if not payload1:
            raise RuntimeError(
                "No encontré el formulario del documento (#asodocumento)."
            )

        url1 = absolute_url(action1) if action1 else resp1.url
        self.log("POST documento ->", url1)
        resp2 = self.post(url1, payload1, referer=resp1.url)

        # Paso 2: password
        soup2 = BeautifulSoup(resp2.text, "html.parser")
        action2, payload2 = self.find_form_payload(soup2, "password", self.password)

        if not payload2:
            # intento extra por si hubo redirect extraño
            resp_retry = self.get(LOGIN_URL)
            soup_retry = BeautifulSoup(resp_retry.text, "html.parser")
            action2, payload2 = self.find_form_payload(
                soup_retry, "password", self.password
            )

        if not payload2:
            raise RuntimeError("No encontré el formulario de password (#password).")

        url2 = absolute_url(action2) if action2 else resp2.url
        self.log("POST password ->", url2)
        resp3 = self.post(url2, payload2, referer=resp2.url)

        # Validación final
        if "tirillaTable" in resp3.text or "/zonatrabajador/tirilla" in resp3.url:
            self.log("Login OK.")
            return

        # Navegación directa a la página final
        resp4 = self.get(TIRILLA_INDEX_URL)
        if "tirillaTable" not in resp4.text:
            preview = clean_text(resp4.text[:500])
            raise RuntimeError(
                "No pude llegar a la página de tirillas tras autenticar. "
                f"URL final: {resp4.url}. Preview: {preview}"
            )

        self.log("Login OK tras navegación manual.")

    def fetch_tirilla_index(self) -> List[TirillaRow]:
        resp = self.get(TIRILLA_INDEX_URL)
        soup = BeautifulSoup(resp.text, "html.parser")

        table = soup.select_one("#tirillaTable")
        if not table:
            raise RuntimeError("No encontré #tirillaTable en la página de tirillas.")

        rows: List[TirillaRow] = []

        for tr in table.select("tbody tr"):
            tds = tr.find_all("td")
            if len(tds) < 3:
                continue

            try:
                orden = int(clean_text(tds[0].get_text()))
            except Exception:
                orden = len(rows) + 1

            fecha_pago = clean_text(tds[1].get_text())
            mes = fecha_pago[:7] if len(fecha_pago) >= 7 else fecha_pago

            btn = tds[2].select_one("button.loadver[request]")
            request_path = btn.get("request", "") if btn else ""

            match = re.search(r"/id/(\d+)", request_path or "")
            id_tirilla = match.group(1) if match else ""

            rows.append(
                TirillaRow(
                    orden=orden,
                    fecha_pago=fecha_pago,
                    mes=mes,
                    request_path=request_path,
                    id_tirilla=id_tirilla,
                    fuente="scraping index",
                )
            )

        return rows

    def extract_label_value(
        self, soup: BeautifulSoup, label_regex: str
    ) -> Optional[str]:
        patt = re.compile(label_regex, re.I)

        for td in soup.find_all("td"):
            text = clean_text(td.get_text(" ", strip=True))
            if patt.search(text):
                sib = td.find_next_sibling("td")
                if sib:
                    return clean_text(sib.get_text(" ", strip=True))

        return None

    def fetch_detail(self, row: TirillaRow) -> TirillaRow:
        if not row.request_path:
            return row

        url = absolute_url(row.request_path)
        resp = self.get(url)
        soup = BeautifulSoup(resp.text, "html.parser")

        row.periodo_pago = self.extract_label_value(soup, r"Per[ií]odo de pago")
        salario = self.extract_label_value(soup, r"Salario Basico")
        row.salario_basico = money_to_int(salario or "")

        # TOTAL A PAGAR
        for td in soup.find_all("td"):
            text = clean_text(td.get_text(" ", strip=True))
            if re.search(r"TOTAL A PAGAR", text, re.I):
                sib = td.find_next_sibling("td")
                if sib:
                    row.total_a_pagar = money_to_int(sib.get_text(" ", strip=True))
                break

        # Totales de tabla
        cells = [clean_text(td.get_text(" ", strip=True)) for td in soup.find_all("td")]
        money_cells = [c for c in cells if re.fullmatch(r"\$[\d,\.]+", c)]

        # Normalmente al final aparecen:
        # total devengos, total deducciones, total a pagar
        if len(money_cells) >= 3:
            row.total_devengos = money_to_int(money_cells[-3])
            row.total_deducciones = money_to_int(money_cells[-2])
            if row.total_a_pagar is None:
                row.total_a_pagar = money_to_int(money_cells[-1])

        row.fuente = "scraping detalle"
        return row

    def scrape(self) -> List[TirillaRow]:
        self.ensure_logged_in()

        rows = self.fetch_tirilla_index()
        if not rows:
            raise RuntimeError("No encontré tirillas.")

        result: List[TirillaRow] = []

        for i, row in enumerate(rows, start=1):
            self.log(f"[{i}/{len(rows)}] {row.fecha_pago} -> {row.request_path}")
            try:
                result.append(self.fetch_detail(row))
            except Exception as exc:
                row.fuente = f"error detalle: {exc}"
                result.append(row)

        return result


def write_excel(rows: List[TirillaRow], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tirillas_detalle"

    headers = [
        "Orden",
        "Fecha de pago",
        "Mes",
        "Periodo de pago",
        "ID tirilla",
        "Salario básico",
        "Total devengos",
        "Total deducciones",
        "Total a pagar",
        "Request path",
        "Fuente",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    for row in rows:
        ws.append(
            [
                row.orden,
                row.fecha_pago,
                row.mes,
                row.periodo_pago,
                row.id_tirilla,
                row.salario_basico,
                row.total_devengos,
                row.total_deducciones,
                row.total_a_pagar,
                row.request_path,
                row.fuente,
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    for col in [6, 7, 8, 9]:
        for cells in ws.iter_cols(
            min_col=col, max_col=col, min_row=2, max_row=ws.max_row
        ):
            for cell in cells:
                format_currency(cell)

    widths = {
        1: 8,
        2: 16,
        3: 12,
        4: 26,
        5: 12,
        6: 16,
        7: 16,
        8: 18,
        9: 16,
        10: 38,
        11: 28,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Resumen mensual
    summary = wb.create_sheet("Resumen_mensual")
    summary_headers = ["Mes", "Total pagado del mes", "Número de tirillas"]
    summary.append(summary_headers)

    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    month_map: Dict[str, List[TirillaRow]] = {}
    for row in rows:
        month_map.setdefault(row.mes, []).append(row)

    for mes in sorted(month_map.keys()):
        tirillas = month_map[mes]
        total_mes = sum(item.total_a_pagar or 0 for item in tirillas)
        summary.append([mes, total_mes, len(tirillas)])

    for cell in summary["B"][1:]:
        format_currency(cell)

    for idx, width in {1: 12, 2: 22, 3: 18}.items():
        summary.column_dimensions[get_column_letter(idx)].width = width

    summary.freeze_panes = "A2"
    summary.auto_filter.ref = f"A1:C{summary.max_row}"

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Scraper de tirillas Zing/Caudata.")
    parser.add_argument(
        "--document", default=CAUDATA_DOCUMENT, help="Documento del trabajador"
    )
    parser.add_argument(
        "--password", default=CAUDATA_PASSWORD, help="Password del trabajador"
    )
    parser.add_argument(
        "--output", default="caudata_pagos.xlsx", help="Archivo Excel de salida"
    )
    parser.add_argument("--timeout", type=int, default=30, help="Timeout por request")
    parser.add_argument("--verbose", action="store_true", help="Mostrar logs")

    args = parser.parse_args()

    scraper = ZingScraper(
        document=args.document,
        password=args.password,
        timeout=args.timeout,
        verbose=args.verbose,
    )

    rows = scraper.scrape()
    write_excel(rows, args.output)

    print(f"OK: archivo generado -> {args.output}")
    print(f"Tirillas procesadas: {len(rows)}")
    print(f"Motor HTTP: {'curl_cffi' if USING_CURL_CFFI else 'requests'}")


if __name__ == "__main__":
    main()
